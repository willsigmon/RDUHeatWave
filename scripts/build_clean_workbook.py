#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import urllib.request
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import ColumnDimension

DEFAULT_HEADERS = [
    'Timestamp', 'Meeting Date', 'First name', 'Last name', 'Profession', 'Company Name',
    'Email', 'Phone', 'Guest of', 'First Visit?', 'Interested in Learning More?',
    'Best Contact Method', 'Ideal Referral'
]

TITLE_FILL = PatternFill('solid', fgColor='E8580C')
HEADER_FILL = PatternFill('solid', fgColor='F4B183')
SUBHEADER_FILL = PatternFill('solid', fgColor='FCE4D6')
HELPER_FILL = PatternFill('solid', fgColor='FFF2CC')
ERROR_FILL = PatternFill('solid', fgColor='FDE9E7')
NONZERO_FILL = PatternFill('solid', fgColor='FFF4E5')
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
MEMBER_ORDER = [
    'Carter Helms', 'Chad Haywood', 'Craig Morrill', 'Dana Walsh', 'Nate Senn',
    'Rusty Sutton', 'Robert Courts', 'Roni Payne', 'Will Sigmon'
]
SHORT_MEMBER_NAMES = {
    'Carter Helms': 'Carter',
    'Chad Haywood': 'Chad',
    'Craig Morrill': 'Craig',
    'Dana Walsh': 'Dana',
    'Nate Senn': 'Nate',
    'Rusty Sutton': 'Rusty',
    'Robert Courts': 'Robert',
    'Roni Payne': 'Roni',
    'Will Sigmon': 'Will',
}
ALIAS_MAP = {
    'cater helms': 'Carter Helms',
    'carter helms': 'Carter Helms',
    'carter': 'Carter Helms',
    'chad hayword': 'Chad Haywood',
    'chad haywood': 'Chad Haywood',
    'chad': 'Chad Haywood',
    'craig morrill': 'Craig Morrill',
    'craig': 'Craig Morrill',
    'dana walsh': 'Dana Walsh',
    'dana': 'Dana Walsh',
    'nathan senn': 'Nate Senn',
    'nate senn': 'Nate Senn',
    'nate': 'Nate Senn',
    'nathan': 'Nate Senn',
    'rusty sutton': 'Rusty Sutton',
    'rusty': 'Rusty Sutton',
    'robert court': 'Robert Courts',
    'robert courts': 'Robert Courts',
    'robert': 'Robert Courts',
    'roni payne': 'Roni Payne',
    'roni': 'Roni Payne',
    'will sigmon': 'Will Sigmon',
    'will': 'Will Sigmon',
    'norre barrett (nm)': 'Noreen Barrett (NM)',
    'norre barrett': 'Noreen Barrett (NM)',
}
STATUS_LIST = ['In Progress', 'Approved', 'Inducted', 'Rejected', 'Waitlisted']
DISPOSITION_LIST = ['In-Progress', 'Closed Business', 'Dead']
YES_NO_LIST = ['Yes', 'No']
LEARN_MORE_LIST = ['Yes', 'Maybe', 'No']
CONTACT_METHOD_LIST = ['Text', 'Call', 'Email']

BACKUP_SHEETS = [
    ('BKP Guest Check In', 'Guest Check In'),
    ('BKP Guest Incentive', 'Guest Incentive Report'),
    ('BKP Membership Apps', 'Membership Applications'),
    ('BKP Member Directory', 'Membership Directory'),
    ('BKP Attendance', 'Attendance Report'),
    ('BKP BizChats', 'BizChats Report'),
    ('BKP Gratitude', 'Gratitude Incentives'),
    ('BKP Referral Pipe', 'Referral Pipeline'),
    ('BKP Passed Ref', 'Passed Referral Tracking Report'),
    ('BKP Revenue', 'Revenue Tracking Report'),
    ('BKP Team Admin', 'Team Administrators Report'),
]


@dataclass
class NeedsReviewRow:
    source_sheet: str
    reason: str
    raw_snapshot: str


@dataclass
class ParsedWorkbook:
    guest_check_in_raw: list[list[str]]
    guest_incentive_raw: list[list[str]]
    membership_apps_raw: list[list[str]]
    membership_directory_raw: list[list[str]]
    attendance_raw: list[list[str]]
    bizchats_raw: list[list[str]]
    gratitude_raw: list[list[str]]
    referral_pipeline_raw: list[list[str]]
    passed_referrals_raw: list[list[str]]
    revenue_raw: list[list[str]]
    team_admin_raw: list[list[str]]


def clean_text(value: object) -> str:
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def lower_email(value: str) -> str:
    return clean_text(value).lower()


def normalize_phone(value: str) -> str:
    digits = re.sub(r'\D', '', clean_text(value))
    if not digits:
        return ''
    if len(digits) == 11 and digits.startswith('1'):
        digits = digits[1:]
    if len(digits) == 10:
        return f'{digits[0:3]}-{digits[3:6]}-{digits[6:10]}'
    return digits


def parse_csv_part(part: str) -> list[list[str]]:
    reader = csv.reader(io.StringIO(part))
    rows = [row[1:] for row in reader]
    return rows


def load_export(path: Path) -> ParsedWorkbook:
    parts = path.read_text().split('\f')
    parsed = [parse_csv_part(part) for part in parts]
    return ParsedWorkbook(*parsed)


def try_parse_date(value: str) -> Optional[date]:
    value = clean_text(value)
    if not value:
        return None
    for fmt in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d', '%m/%d/%Y'):
        try:
            dt = datetime.strptime(value, fmt)
            if fmt == '%m/%d':
                return date(2000, dt.month, dt.day)
            return dt.date()
        except ValueError:
            continue
    return None


def format_birthday(value: str) -> str:
    value = clean_text(value)
    if not value:
        return ''
    dt = try_parse_date(value)
    if not dt:
        return value
    return f'{dt.month:02d}/{dt.day:02d}'


def canonicalize_member(value: str) -> str:
    value = clean_text(value)
    if not value:
        return ''
    normalized = re.sub(r'\s*\(via.*?\)$', '', value, flags=re.I).strip()
    normalized = re.sub(r'\s*\(nm\)$', ' (NM)', normalized, flags=re.I)
    key = normalized.lower()
    if key in ALIAS_MAP:
        return ALIAS_MAP[key]
    # Best-effort substring match for canonical members.
    for canonical in MEMBER_ORDER:
        if canonical.lower() in key:
            return canonical
        first = canonical.split()[0].lower()
        if re.fullmatch(fr'{re.escape(first)}\b', key):
            return canonical
    return normalized


def canonicalize_given_party(value: str) -> str:
    value = clean_text(value)
    if not value:
        return ''
    key = value.lower().strip()
    if key in ALIAS_MAP:
        return ALIAS_MAP[key]
    for canonical in MEMBER_ORDER:
        if canonical.lower() == key:
            return canonical
    return value.strip()


def parse_source_rows(rows: list[list[str]], header_row_idx: int) -> tuple[list[str], list[list[str]]]:
    header = [clean_text(cell) for cell in rows[header_row_idx]]
    data_rows = []
    for row in rows[header_row_idx + 1:]:
        trimmed = row[:len(header)] + [''] * max(0, len(header) - len(row))
        if any(clean_text(cell) for cell in trimmed):
            data_rows.append(trimmed[:len(header)])
    return header, data_rows


def build_guest_checkin(parsed: ParsedWorkbook, needs_review: list[NeedsReviewRow], change_log: list[list[str]]):
    header, raw_rows = parse_source_rows(parsed.guest_check_in_raw, 2)
    cleaned_rows = []
    for row in raw_rows:
        row = [clean_text(cell) for cell in row]
        if len(row) < 9:
            row += [''] * (9 - len(row))
        # Broken March 10 shifted row
        if row[0].startswith('2026-03-10') and row[1] == 'Will' and row[2] == 'Sigmon':
            needs_review.append(NeedsReviewRow('Guest Check In', 'Shifted/misaligned web form row quarantined', json.dumps(row)))
            change_log.append([timestamp(), 'Guest Check In', 'Moved row to Needs Review', 'Shifted March 10, 2026 row was quarantined instead of guessed'])
            continue
        normalized = {
            'Timestamp': row[0],
            'Meeting Date': row[1],
            'First name': row[2],
            'Last name': row[3],
            'Profession': row[4],
            'Company Name': '',
            'Email': lower_email(row[5]),
            'Phone': normalize_phone(row[6]),
            'Guest of': canonicalize_member(row[7]),
            'First Visit?': 'Yes' if clean_text(row[8]).lower() == 'yes' else 'No' if clean_text(row[8]).lower() == 'no' else clean_text(row[8]),
            'Interested in Learning More?': '',
            'Best Contact Method': '',
            'Ideal Referral': '',
        }
        if row[7] != normalized['Guest of']:
            change_log.append([timestamp(), 'Guest Check In', 'Normalized Guest of', f"{row[7]} → {normalized['Guest of']}"])
        if row[5] != normalized['Email'] and row[5]:
            change_log.append([timestamp(), 'Guest Check In', 'Normalized email', f"{row[5]} → {normalized['Email']}"])
        if row[6] and row[6] != normalized['Phone']:
            change_log.append([timestamp(), 'Guest Check In', 'Normalized phone', f"{row[6]} → {normalized['Phone']}"])
        cleaned_rows.append([normalized[h] for h in DEFAULT_HEADERS])
    return cleaned_rows


def build_membership_apps(parsed: ParsedWorkbook, needs_review: list[NeedsReviewRow], change_log: list[list[str]]):
    header, raw_rows = parse_source_rows(parsed.membership_apps_raw, 3)
    cleaned = []
    for row in raw_rows:
        values = [clean_text(cell) for cell in row]
        if values[0].startswith('2026-03-09') and values[1] in {'Verify', 'Formy'}:
            needs_review.append(NeedsReviewRow('Membership Applications', 'Test/form row quarantined', json.dumps(values)))
            change_log.append([timestamp(), 'Membership Applications', 'Moved row to Needs Review', f'Test row for {values[1]} quarantined'])
            continue
        record = values[:]
        record[2] = normalize_phone(record[2])
        record[3] = lower_email(record[3])
        if len(record) > 4:
            original = record[4]
            record[4] = canonicalize_member(record[4])
            if original != record[4]:
                change_log.append([timestamp(), 'Membership Applications', 'Normalized sponsor', f'{original} → {record[4]}'])
        if len(record) > 5:
            record[5] = canonicalize_member(record[5])
        if len(record) > 7 and record[7]:
            status = record[7].strip()
            mapped = {
                'in progress': 'In Progress',
                'approved': 'Approved',
                'inducted': 'Inducted',
            }.get(status.lower(), status)
            record[7] = mapped
        cleaned.append(record)
    return header, cleaned


def build_membership_directory(parsed: ParsedWorkbook, change_log: list[list[str]]):
    header, raw_rows = parse_source_rows(parsed.membership_directory_raw, 3)
    cleaned = []
    for row in raw_rows:
        record = [clean_text(cell) for cell in row]
        record[4] = normalize_phone(record[4])
        record[5] = lower_email(record[5])
        original_bday = record[6]
        record[6] = format_birthday(record[6])
        if original_bday != record[6] and original_bday:
            change_log.append([timestamp(), 'Membership Directory', 'Standardized birthday', f'{original_bday} → {record[6]}'])
        if record[3] == 'Mortgage Loan Offier':
            record[3] = 'Mortgage Loan Officer'
            change_log.append([timestamp(), 'Membership Directory', 'Fixed typo', 'Mortgage Loan Offier → Mortgage Loan Officer'])
        cleaned.append(record)
    return header, cleaned


def build_referral_pipeline(parsed: ParsedWorkbook, change_log: list[list[str]]):
    header, raw_rows = parse_source_rows(parsed.referral_pipeline_raw, 3)
    cleaned = []
    for row in raw_rows:
        record = [clean_text(cell) for cell in row]
        record[0] = canonicalize_given_party(record[0])
        record[1] = canonicalize_member(record[1]) or clean_text(row[1])
        if record[2] == '2025-02-19 00:00:00':
            record[2] = '2026-02-19 00:00:00'
            change_log.append([timestamp(), 'Referral Pipeline', 'Fixed year', f"{row[2]} → {record[2]}"])
        if record[5] in {'', None}:
            record[5] = ''
        else:
            try:
                revenue = float(str(record[5]).replace(',', ''))
                record[5] = int(revenue) if revenue.is_integer() else revenue
            except Exception:
                pass
        cleaned.append(record)
    return header, cleaned


def timestamp() -> str:
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def write_table(ws, rows: list[list[object]], title_row: int | None = None, header_row: int | None = None):
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = LEFT
    if title_row:
        for cell in ws[title_row]:
            if cell.value not in (None, ''):
                cell.fill = TITLE_FILL
                cell.font = Font(color='FFFFFF', bold=True, size=12)
                cell.alignment = CENTER
    if header_row:
        for cell in ws[header_row]:
            if cell.value not in (None, ''):
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True)
                cell.alignment = CENTER
    autosize(ws)


def autosize(ws):
    widths = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths[cell.column], len(str(cell.value)))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 32)


def style_source_sheet(ws, title: str, headers: list[str], rows: list[list[object]], freeze_row: int, filter_row: int):
    ws.append([title])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    write_table(ws, [], None, None)
    for cell in ws[1]:
        if cell.value:
            cell.fill = TITLE_FILL
            cell.font = Font(color='FFFFFF', bold=True, size=12)
            cell.alignment = LEFT
    for cell in ws[2]:
        if cell.value:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.alignment = CENTER
            cell.border = BORDER
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.border = BORDER
            cell.alignment = LEFT
    ws.freeze_panes = f'A{freeze_row}'
    ws.auto_filter.ref = f'A{filter_row}:{get_column_letter(ws.max_column)}{ws.max_row}'
    autosize(ws)


def add_dropdown(ws, cell_range: str, formula: str):
    dv = DataValidation(type='list', formula1=formula, allow_blank=True)
    dv.error = 'Select a value from the list.'
    dv.prompt = 'Choose from the dropdown list.'
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_incomplete_rule(ws, start_row: int, end_col: int, formula_cols: list[int]):
    col_letters = [get_column_letter(c) for c in formula_cols]
    checks = ','.join(f'${col}{start_row}=""' for col in col_letters)
    formula = f'=OR({checks})'
    rule = FormulaRule(formula=[formula], fill=ERROR_FILL)
    ws.conditional_formatting.add(f'A{start_row}:{get_column_letter(end_col)}{ws.max_row}', rule)


def create_validation_sheet(wb: Workbook):
    ws = wb.create_sheet('Validation Lists')
    sections = {
        'Members': MEMBER_ORDER,
        'Application Status': STATUS_LIST,
        'Disposition': DISPOSITION_LIST,
        'Yes / No': YES_NO_LIST,
        'Learn More': LEARN_MORE_LIST,
        'Best Contact': CONTACT_METHOD_LIST,
    }
    row = 1
    for title, values in sections.items():
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=1).font = Font(bold=True)
        for idx, value in enumerate(values, start=row + 1):
            ws.cell(row=idx, column=1, value=value)
        row += len(values) + 3
    autosize(ws)
    return ws


def create_alias_sheet(wb: Workbook):
    ws = wb.create_sheet('Alias Map')
    ws.append(['Alias', 'Canonical Name'])
    for alias, canonical in sorted(ALIAS_MAP.items()):
        ws.append([alias, canonical])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    autosize(ws)
    return ws


def create_change_log(wb: Workbook, change_log: list[list[str]]):
    ws = wb.create_sheet('Change Log')
    ws.append(['Timestamp', 'Sheet', 'Action', 'Detail'])
    for row in change_log:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:D{ws.max_row}'
    autosize(ws)
    return ws


def create_needs_review(wb: Workbook, rows: list[NeedsReviewRow]):
    ws = wb.create_sheet('Needs Review')
    ws.append(['Source Sheet', 'Reason', 'Raw Snapshot'])
    for item in rows:
        ws.append([item.source_sheet, item.reason, item.raw_snapshot])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:C{ws.max_row}'
    ws.column_dimensions['C'].width = 90
    autosize(ws)
    return ws


def set_sheet_formats(ws, date_cols=(), datetime_cols=(), currency_cols=(), int_cols=(), text_cols=()):
    for col in date_cols:
        for cell in ws[col][2:]:
            if cell.row >= 3:
                cell.number_format = 'm/d/yyyy'
    for col in datetime_cols:
        for cell in ws[col][2:]:
            if cell.row >= 3:
                cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'
    for col in currency_cols:
        for cell in ws[col][1:]:
            cell.number_format = '$#,##0'
    for col in int_cols:
        for cell in ws[col][1:]:
            cell.number_format = '0'
    for col in text_cols:
        for cell in ws[col]:
            cell.number_format = '@'


def parse_week_dates(team_admin_raw: list[list[str]]) -> list[str]:
    dates = []
    for row in team_admin_raw[3:54]:
        if row and clean_text(row[0]):
            dates.append(clean_text(row[0]))
    return dates


def create_passed_referrals(ws, weeks: list[str]):
    title = 'PASSED REFERRAL TRACKING REPORT'
    ws.append([title])
    header1 = ['Rolling 12 Mo']
    header2 = ['']
    for member in MEMBER_ORDER:
        header1.extend([SHORT_MEMBER_NAMES[member], ''])
        header2.extend(['Given', 'Rcvd'])
    header1.extend(['Weekly Total', ''])
    header2.extend(['Given', 'Rcvd'])
    ws.append(header1)
    ws.append(header2)
    for idx, week in enumerate(weeks, start=4):
        ws.cell(idx, 1, week)
        for member_index, member in enumerate(MEMBER_ORDER):
            col = 2 + member_index * 2
            ws.cell(idx, col, f'=COUNTIFS(\'Referral Pipeline\'!$A:$A,{json.dumps(member)},\'Referral Pipeline\'!$C:$C,$A{idx})')
            ws.cell(idx, col + 1, f'=COUNTIFS(\'Referral Pipeline\'!$B:$B,{json.dumps(member)},\'Referral Pipeline\'!$C:$C,$A{idx})')
        ws.cell(idx, 20, f'=COUNTIF(\'Referral Pipeline\'!$C:$C,$A{idx})')
        ws.cell(idx, 21, f'=COUNTIF(\'Referral Pipeline\'!$C:$C,$A{idx})')
    total_row = len(weeks) + 4
    ws.cell(total_row, 1, 'TOTAL')
    for col in range(2, 22):
        letter = get_column_letter(col)
        ws.cell(total_row, col, f'=SUM({letter}4:{letter}{total_row-1})')
    format_report_sheet(ws, title_row=1, header_rows=(2, 3), freeze='A4', currency_cols=(), hidden_start='2026-03-12 00:00:00')


def create_revenue_tracking(ws, weeks: list[str]):
    title = 'REVENUE TRACKING REPORT'
    ws.append([title])
    header1 = ['Rolling 12 Mo']
    header2 = ['']
    for member in MEMBER_ORDER:
        header1.extend([SHORT_MEMBER_NAMES[member], ''])
        header2.extend(['Given', 'Rcvd'])
    header1.extend(['Weekly Total', ''])
    header2.extend(['Given', 'Rcvd'])
    ws.append(header1)
    ws.append(header2)
    for idx, week in enumerate(weeks, start=4):
        ws.cell(idx, 1, week)
        for member_index, member in enumerate(MEMBER_ORDER):
            col = 2 + member_index * 2
            ws.cell(idx, col, (
                f'=SUMIFS(\'Referral Pipeline\'!$F:$F,\'Referral Pipeline\'!$A:$A,{json.dumps(member)},'
                f'\'Referral Pipeline\'!$C:$C,$A{idx},\'Referral Pipeline\'!$E:$E,"Closed Business")'
            ))
            ws.cell(idx, col + 1, (
                f'=SUMIFS(\'Referral Pipeline\'!$F:$F,\'Referral Pipeline\'!$B:$B,{json.dumps(member)},'
                f'\'Referral Pipeline\'!$C:$C,$A{idx},\'Referral Pipeline\'!$E:$E,"Closed Business")'
            ))
        total_formula = f'=SUMIFS(\'Referral Pipeline\'!$F:$F,\'Referral Pipeline\'!$C:$C,$A{idx},\'Referral Pipeline\'!$E:$E,"Closed Business")'
        ws.cell(idx, 20, total_formula)
        ws.cell(idx, 21, total_formula)
    total_row = len(weeks) + 4
    ws.cell(total_row, 1, 'TOTAL')
    for col in range(2, 22):
        letter = get_column_letter(col)
        ws.cell(total_row, col, f'=SUM({letter}4:{letter}{total_row-1})')
    format_report_sheet(ws, title_row=1, header_rows=(2, 3), freeze='A4', currency_cols=tuple(get_column_letter(i) for i in range(2, 22)), hidden_start='2026-03-12 00:00:00')


def create_team_admin(ws, weeks: list[str]):
    title = 'TEAM ADMINISTRATORS REPORT'
    ws.append([title])
    ws.append(['Rolling 12 Mo', 'BizChats', 'Corporate GIs', 'Member GIs', 'Guests', 'Received Referrals', 'Revenue'])
    for idx, week in enumerate(weeks, start=3):
        ws.cell(idx, 1, week)
        ws.cell(idx, 2, f'=INDEX(\'BizChats Report\'!$K$3:$K$55,MATCH($A{idx},\'BizChats Report\'!$A$3:$A$55,0))')
        ws.cell(idx, 3, f'=INDEX(\'Gratitude Incentives\'!$T$60:$T$112,MATCH($A{idx},\'Gratitude Incentives\'!$A$60:$A$112,0))')
        ws.cell(idx, 4, f'=INDEX(\'Gratitude Incentives\'!$U$4:$U$56,MATCH($A{idx},\'Gratitude Incentives\'!$A$4:$A$56,0))')
        ws.cell(idx, 5, f'=INDEX(\'Guest Incentive Report\'!$R$4:$R$55,MATCH($A{idx},\'Guest Incentive Report\'!$A$4:$A$55,0))')
        ws.cell(idx, 6, f'=COUNTIF(\'Referral Pipeline\'!$C:$C,$A{idx})')
        ws.cell(idx, 7, f'=SUMIFS(\'Referral Pipeline\'!$F:$F,\'Referral Pipeline\'!$C:$C,$A{idx},\'Referral Pipeline\'!$E:$E,"Closed Business")')
    label_row = len(weeks) + 3
    total_row = len(weeks) + 4
    ws.cell(label_row, 1, 'Rolling 12 Mo')
    ws.cell(label_row, 2, 'BizChats')
    ws.cell(label_row, 3, 'CGIs')
    ws.cell(label_row, 4, 'MGIs')
    ws.cell(label_row, 5, 'Guests')
    ws.cell(label_row, 6, 'Referrals')
    ws.cell(label_row, 7, 'Revenue')
    ws.cell(total_row, 1, 'TOTAL')
    for col in range(2, 8):
        letter = get_column_letter(col)
        ws.cell(total_row, col, f'=SUM({letter}3:{letter}{label_row-1})')
    format_report_sheet(ws, title_row=1, header_rows=(2,), freeze='A3', currency_cols=('G',), hidden_start='2026-03-12 00:00:00')
    ws.row_dimensions[label_row].fill = HELPER_FILL


def create_revenue_reconciliation(ws, weeks: list[str]):
    ws.append(['Week', 'Pipeline Closed Revenue', 'Revenue Report Total', 'Delta', 'Closed Deals', 'Report Given Total', 'Report Received Total'])
    for idx, week in enumerate(weeks, start=2):
        ws.cell(idx, 1, week)
        ws.cell(idx, 2, f'=SUMIFS(\'Referral Pipeline\'!$F:$F,\'Referral Pipeline\'!$C:$C,$A{idx},\'Referral Pipeline\'!$E:$E,"Closed Business")')
        rev_row = idx + 2
        ws.cell(idx, 3, f"='Revenue Tracking Report'!$U${rev_row}")
        ws.cell(idx, 4, f'=B{idx}-C{idx}')
        ws.cell(idx, 5, f'=COUNTIFS(\'Referral Pipeline\'!$C:$C,$A{idx},\'Referral Pipeline\'!$E:$E,"Closed Business")')
        ws.cell(idx, 6, f"='Revenue Tracking Report'!$T${rev_row}")
        ws.cell(idx, 7, f"='Revenue Tracking Report'!$U${rev_row}")
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:G{ws.max_row}'
    for col in ('B', 'C', 'D', 'F', 'G'):
        for cell in ws[col][1:]:
            cell.number_format = '$#,##0'
    autosize(ws)


def format_report_sheet(ws, title_row=1, header_rows=(2,), freeze='A2', currency_cols=(), hidden_start: str | None = None):
    for cell in ws[title_row]:
        if cell.value:
            cell.fill = TITLE_FILL
            cell.font = Font(color='FFFFFF', bold=True, size=12)
            cell.alignment = CENTER
    for row_num in header_rows:
        for cell in ws[row_num]:
            if cell.value not in ('', None):
                cell.fill = HEADER_FILL if row_num == header_rows[0] else SUBHEADER_FILL
                cell.font = Font(bold=True)
                cell.alignment = CENTER
    for row in ws.iter_rows(min_row=max(header_rows)+1, max_row=ws.max_row):
        for cell in row:
            cell.border = BORDER
            if cell.column == 1:
                cell.number_format = 'm/d/yyyy'
                cell.alignment = LEFT
            else:
                cell.alignment = CENTER
    for col in currency_cols:
        for cell in ws[col][max(header_rows):]:
            cell.number_format = '$#,##0'
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f'A{header_rows[-1]}:{get_column_letter(ws.max_column)}{ws.max_row}'
    autosize(ws)
    if hidden_start:
        started = False
        for row in range(max(header_rows)+1, ws.max_row + 1):
            if ws.cell(row, 1).value == hidden_start:
                started = True
            if started and ws.cell(row, 1).value not in {'TOTAL', 'Totals YTD', 'Rolling 12 Mo', None, ''}:
                numeric_values = [ws.cell(row, col).value for col in range(2, ws.max_column + 1)]
                if not any(v not in ('', None, 0, '0') for v in numeric_values):
                    ws.row_dimensions[row].hidden = True
                    ws.row_dimensions[row].outlineLevel = 1


def import_raw_sheet(wb: Workbook, backup_name: str, raw_rows: list[list[str]]):
    ws = wb.create_sheet(backup_name)
    for row in raw_rows:
        ws.append([clean_text(c) for c in row])
    # Style first three rows lightly if present.
    for cell in ws[1]:
        if cell.value:
            cell.fill = TITLE_FILL
            cell.font = Font(color='FFFFFF', bold=True)
    for cell in ws[2]:
        if cell.value:
            cell.fill = SUBHEADER_FILL
            cell.font = Font(bold=True)
    for cell in ws[3]:
        if cell.value:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
    autosize(ws)


def populate_imported_report(
    ws,
    raw_rows: list[list[str]],
    freeze: str,
    currency_cols=(),
    hide_helper_rows: Iterable[int] = (),
    header_rows: tuple[int, ...] = (2, 3),
):
    normalized_rows = raw_rows[:]
    if normalized_rows and all(clean_text(cell).startswith('Unnamed:') for cell in normalized_rows[0] if clean_text(cell)):
        normalized_rows = normalized_rows[1:]

    for row in normalized_rows:
        cleaned = [None if clean_text(c) == '' else (None if clean_text(c) == '-' else clean_text(c)) for c in row]
        ws.append(cleaned)

    title = clean_text(normalized_rows[0][0]) if normalized_rows else ws.title.upper()
    if ws.max_row >= 1 and ws.max_column >= 1 and ws['A1'].value in (None, ''):
        ws['A1'] = title
    if ws['A1'].value:
        for cell in ws[1]:
            if cell.value:
                cell.fill = TITLE_FILL
                cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = CENTER

    first_header_row = header_rows[0] if header_rows else 2
    for row_num in header_rows:
        if ws.max_row >= row_num:
            fill = HEADER_FILL if row_num == first_header_row else SUBHEADER_FILL
            for cell in ws[row_num]:
                if cell.value not in (None, ''):
                    cell.fill = fill
                    cell.font = Font(bold=True)
                    cell.alignment = CENTER

    data_start_row = max(header_rows) + 1 if header_rows else 2
    for row in ws.iter_rows(min_row=data_start_row):
        for cell in row:
            cell.border = BORDER
            cell.alignment = CENTER if cell.column != 1 else LEFT
    ws.freeze_panes = freeze
    filter_row = max(header_rows) if header_rows else 1
    ws.auto_filter.ref = f'A{filter_row}:{get_column_letter(ws.max_column)}{ws.max_row}'
    for col in currency_cols:
        for cell in ws[col][data_start_row - 1:]:
            cell.number_format = '$#,##0'
    autosize(ws)
    for r in hide_helper_rows:
        ws.row_dimensions[r].hidden = True


def upload_to_drive(xlsx_path: Path, file_name: str, parent_folder: Optional[str] = None) -> dict:
    creds = json.loads((Path.home() / '.clasprc.json').read_text())['tokens']['default']
    token = creds['access_token']
    boundary = 'rdu-clean-upload-boundary'
    metadata = {'name': file_name, 'mimeType': 'application/vnd.google-apps.spreadsheet'}
    if parent_folder:
        metadata['parents'] = [parent_folder]
    xlsx_data = xlsx_path.read_bytes()
    body = []
    body.append(f'--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n'.encode())
    body.append(json.dumps(metadata).encode())
    body.append(f'\r\n--{boundary}\r\nContent-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n'.encode())
    body.append(xlsx_data)
    body.append(f'\r\n--{boundary}--'.encode())
    payload = b''.join(body)
    req = urllib.request.Request(
        'https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart&supportsAllDrives=true',
        data=payload,
        method='POST',
        headers={
            'Authorization': 'Bearer ' + token,
            'Content-Type': f'multipart/related; boundary={boundary}',
        },
    )
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read().decode())


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', type=Path, required=True)
    parser.add_argument('--output', type=Path, required=True)
    parser.add_argument('--upload-google-sheet', action='store_true')
    parser.add_argument('--uploaded-name', default='RDU Heatwave Team Stats 2026 - Cleaned')
    args = parser.parse_args()

    parsed = load_export(args.input)
    wb = Workbook()
    wb.remove(wb.active)

    change_log: list[list[str]] = [
        [timestamp(), 'Workbook', 'Rebuilt from export', 'Created a reversible cleaned workbook from the live sheet export because direct in-place Google mutation was blocked in this environment'],
        [timestamp(), 'Workbook', 'Added helper tabs', 'Added Change Log, Validation Lists, Alias Map, Needs Review, and Revenue Reconciliation'],
    ]
    needs_review: list[NeedsReviewRow] = []

    guest_rows = build_guest_checkin(parsed, needs_review, change_log)
    apps_header, apps_rows = build_membership_apps(parsed, needs_review, change_log)
    dir_header, dir_rows = build_membership_directory(parsed, change_log)
    pipe_header, pipe_rows = build_referral_pipeline(parsed, change_log)
    weeks = parse_week_dates(parsed.team_admin_raw)

    ws = wb.create_sheet('Guest Check In')
    style_source_sheet(ws, 'Guest Check In', DEFAULT_HEADERS, guest_rows, freeze_row=3, filter_row=2)
    set_sheet_formats(ws, date_cols=('B',), datetime_cols=('A',), text_cols=('G', 'H', 'I', 'J', 'K', 'L', 'M'))
    add_dropdown(ws, f'I3:I{max(ws.max_row, 500)}', "='Validation Lists'!$A$2:$A$10")
    add_dropdown(ws, f'J3:J{max(ws.max_row, 500)}', '="Yes,No"')
    add_dropdown(ws, f'K3:K{max(ws.max_row, 500)}', '="Yes,Maybe,No"')
    add_dropdown(ws, f'L3:L{max(ws.max_row, 500)}', '="Text,Call,Email"')
    add_incomplete_rule(ws, 3, ws.max_column, [3,4,5,7,8,9,10])

    ws = wb.create_sheet('Membership Applications')
    style_source_sheet(ws, 'Membership Applications', apps_header, apps_rows, freeze_row=3, filter_row=2)
    set_sheet_formats(ws, text_cols=('C', 'D'))
    add_dropdown(ws, f'F3:F{max(ws.max_row, 200)}', "='Validation Lists'!$A$2:$A$10")
    add_dropdown(ws, f'G3:G{max(ws.max_row, 200)}', "='Validation Lists'!$A$2:$A$10")
    add_dropdown(ws, f'H3:H{max(ws.max_row, 200)}', "='Validation Lists'!$A$2:$A$10")
    add_dropdown(ws, f'I3:I{max(ws.max_row, 200)}', '="In Progress,Approved,Inducted,Rejected,Waitlisted"')
    add_incomplete_rule(ws, 3, ws.max_column, [1,2,3,4,5,8])

    ws = wb.create_sheet('Membership Directory')
    style_source_sheet(ws, 'Membership Directory', dir_header, dir_rows, freeze_row=3, filter_row=2)
    set_sheet_formats(ws, text_cols=('E', 'F', 'G'))
    add_incomplete_rule(ws, 3, ws.max_column, [1,2,3,4,5,6])

    ws = wb.create_sheet('Referral Pipeline')
    style_source_sheet(ws, 'Referral Pipeline', pipe_header, pipe_rows, freeze_row=3, filter_row=2)
    set_sheet_formats(ws, date_cols=('C',), currency_cols=('F',), text_cols=('A','B','D','E'))
    add_dropdown(ws, f'B3:B{max(ws.max_row, 300)}', "='Validation Lists'!$A$2:$A$10")
    add_dropdown(ws, f'E3:E{max(ws.max_row, 300)}', '="In-Progress,Closed Business,Dead"')
    add_incomplete_rule(ws, 3, ws.max_column, [1,2,3,4,5])

    ws = wb.create_sheet('Guest Incentive Report')
    populate_imported_report(ws, parsed.guest_incentive_raw, freeze='A4', hide_helper_rows=(55,), header_rows=(2, 3))

    ws = wb.create_sheet('Attendance Report')
    populate_imported_report(ws, parsed.attendance_raw, freeze='A4', header_rows=(2, 3))

    ws = wb.create_sheet('BizChats Report')
    populate_imported_report(ws, parsed.bizchats_raw, freeze='A3', header_rows=(2,))

    ws = wb.create_sheet('Gratitude Incentives')
    populate_imported_report(ws, parsed.gratitude_raw, freeze='A4', header_rows=(2, 3))

    ws = wb.create_sheet('Passed Referral Tracking Report')
    create_passed_referrals(ws, weeks)

    ws = wb.create_sheet('Revenue Tracking Report')
    create_revenue_tracking(ws, weeks)

    ws = wb.create_sheet('Team Administrators Report')
    create_team_admin(ws, weeks)

    create_validation_sheet(wb)
    create_alias_sheet(wb)
    create_needs_review(wb, needs_review)
    create_revenue_reconciliation(wb.create_sheet('Revenue Reconciliation'), weeks)
    create_change_log(wb, change_log)

    for backup_name, original_name in BACKUP_SHEETS:
        mapping = {
            'Guest Check In': parsed.guest_check_in_raw,
            'Guest Incentive Report': parsed.guest_incentive_raw,
            'Membership Applications': parsed.membership_apps_raw,
            'Membership Directory': parsed.membership_directory_raw,
            'Attendance Report': parsed.attendance_raw,
            'BizChats Report': parsed.bizchats_raw,
            'Gratitude Incentives': parsed.gratitude_raw,
            'Referral Pipeline': parsed.referral_pipeline_raw,
            'Passed Referral Tracking Report': parsed.passed_referrals_raw,
            'Revenue Tracking Report': parsed.revenue_raw,
            'Team Administrators Report': parsed.team_admin_raw,
        }
        import_raw_sheet(wb, backup_name, mapping[original_name])

    # Protect formula/report tabs.
    for sheet_name in ['Guest Incentive Report', 'Attendance Report', 'BizChats Report', 'Gratitude Incentives', 'Passed Referral Tracking Report', 'Revenue Tracking Report', 'Team Administrators Report', 'Revenue Reconciliation']:
        ws = wb[sheet_name]
        ws.protection.sheet = True
        ws.protection.enable()
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f'Wrote workbook to {args.output}')

    if args.upload_google_sheet:
        result = upload_to_drive(args.output, args.uploaded_name)
        print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()
